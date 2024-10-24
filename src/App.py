import os
import sys
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
import shutil
import datetime
import tkinter as tk
from tkinter import messagebox, filedialog, ttk
from tkcalendar import DateEntry


class CancelamentoSelecao(Exception):
    pass
def selecionar_arquivo():
    caminho_arquivo = filedialog.askopenfilename(
        title="Selecione a Planilha",
        filetypes=[("Arquivos Excel", "*.xlsx"), ("Todos os Arquivos", "*.*")]
    )
    if caminho_arquivo:
        return caminho_arquivo
    else:
        raise CancelamentoSelecao()
def finalizar():
    sys.exit()

root = tk.Tk()
root.withdraw()

try:
# Uso
    PATH_ARQUIVO = selecionar_arquivo()

except CancelamentoSelecao:
    finalizar()

def configurar_formatacao(ws):
    alinhar = Alignment(horizontal='center', vertical='center')

    #Definir a largura das colunas e aplicar alinhamento a todas as células
    for col in ws.columns:
        col_letter = col[0].column_letter
        ws.column_dimensions[col_letter].width = 22  #Definindo a largura da coluna

    #Aplicar alinhamento e formatação de células em todas as linhas
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = alinhar
        #Formatar a primeira coluna como data e a segunda como número com separadores
        row[0].number_format = 'MM/DD/YYYY'  #Coluna de DATA
        row[1].number_format = '#,##0.00'  #Coluna de VALOR

def aplicar_estilo_cabecalho(ws):
    cor_cabecalho = PatternFill(start_color='9BC2E6', end_color='9BC2E6', fill_type="solid")
    alinhar = Alignment(horizontal='center', vertical='center')

    #Aplicar estilo no cabeçalho (primeira linha)
    for cell in ws[1]:
        cell.fill = cor_cabecalho
        cell.alignment = alinhar

def copiar_cabecalho(ws_origem, ws_destino):
    #Copia o cabeçalho e as larguras de colunas da aba 'Financeiro Geral' para a aba do centro de custo.
    for i, cell in enumerate(ws_origem[1], 1):
        ws_destino.cell(row=1, column=i).value = cell.value
        ws_destino.cell(row=1, column=i).alignment = Alignment(horizontal='center', vertical='center')

    #Copiar a largura das colunas
    for col in ws_origem.columns:
        col_letter = col[0].column_letter
        ws_destino.column_dimensions[col_letter].width = ws_origem.column_dimensions[col_letter].width

    aplicar_estilo_cabecalho(ws_destino)

def carregar_planilha(path_arquivo, sheet_name='Financeiro Geral'):
    if os.path.exists(path_arquivo):
        df = pd.read_excel(path_arquivo, sheet_name=sheet_name)
        df['VALOR'] = pd.to_numeric(df['VALOR'], errors='coerce')
        df['DATA'] = pd.to_datetime(df['DATA'], errors='coerce')
        return df
    
    else:
        raise FileNotFoundError(f"Arquivo {path_arquivo} não encontrado.")

def filtrar_por_centro(df, centro):
    return df[df['CENTRO'] == centro]

def aba_existe(wb, centro):
    return centro in wb.sheetnames

def obter_dados_aba(ws):
    #Extrai os dados existentes de uma aba como uma lista de tuplas.
    dados_existentes = []
    for row in ws.iter_rows(min_row=2, values_only=True):  #Ignorar o cabeçalho
        dados_existentes.append(row)
    return dados_existentes

def atualizar_ou_criar_aba(path_arquivo, df_filtrado, centro):
    #Atualiza ou cria uma nova aba no arquivo Excel com a mesma formatação da aba original.
    #Apenas adiciona dados novos que ainda não existem na aba.
    
    #Carregar o arquivo Excel existente
    wb = load_workbook(path_arquivo)
    ws_geral = wb['Financeiro Geral']  #Pegamos a aba original para copiar o cabeçalho e formatação
    
    if aba_existe(wb, centro):
        ws = wb[centro]
        dados_existentes = obter_dados_aba(ws)
        #Adicionar apenas os novos dados que não estão na aba existente
        for r in dataframe_to_rows(df_filtrado, index=False, header=False):
            if tuple(r) not in dados_existentes:
                ws.append(r)
    else:
        #Criar uma nova aba para o centro de custo
        ws = wb.create_sheet(title=centro)
        for r in dataframe_to_rows(df_filtrado, index=False, header=True):
            ws.append(r)
        #Copiar o cabeçalho e formatação da aba 'Financeiro Geral'
        copiar_cabecalho(ws_geral, ws)
    
    #Aplicar formatação geral
    configurar_formatacao(ws)
    
    #Salvar o arquivo Excel com as alterações
    wb.save(path_arquivo)

def obter_ultima_linha_vazia(ws):
    #Obter a última linha vazia da aba.
    for row in range(2, ws.max_row + 1):
        if all(cell.value is None for cell in ws[row]):
            return row
    return ws.max_row + 1

def adicionar_dados_financeiro_geral(path_arquivo, novo_dado):
    fazer_backup(path_arquivo)
    df = carregar_planilha(path_arquivo)
    df_novo_dado = pd.DataFrame([novo_dado])
    df_atualizado = pd.concat([df, df_novo_dado], ignore_index=True)
    
    #Salva de volta na aba 'Financeiro Geral' sem sobrescrever as outras
    wb = load_workbook(path_arquivo)
    ws = wb['Financeiro Geral']
    
    try:
        #Encontra a primeira linha vazia
        ultima_vazia=obter_ultima_linha_vazia(ws)

        for i, row in enumerate(dataframe_to_rows(df_novo_dado, index=False, header=False), start=ultima_vazia):
            for j, value in enumerate(row, start=1):
                # Verifica se o valor é um número e converte para string se necessário
                if isinstance(value, (float, int)):
                    ws.cell(row=i, column=j, value=value)  # Mantém como número
                else:
                    ws.cell(row=i, column=j, value=str(value) if value is not None else None)

        aplicar_estilo_cabecalho(ws)  #Aplica o estilo do cabeçalho na aba 'Financeiro Geral'
        configurar_formatacao(ws)  #Formata a planilha
    
    #Salva o arquivo Excel com as alterações
        wb.save(path_arquivo)

    #Atualiza as abas dos centros de custo
        atualizar_abas_centros(path_arquivo)
        
        return True
    
    except PermissionError:
        messagebox.showerror("Erro de Permissão", "A planilha está aberta. Por favor, feche a planilha antes de adicionar novos dados.")
        return False
    

def fazer_backup(path_arquivo, max_backups=3):   #Altere aqui o tanto de backups que voce deseja manter
    pasta_backup = "backups"

    if not os.path.exists(pasta_backup):
        os.makedirs(pasta_backup) 

    #Cria um backup do arquivo Excel existente
    timestamp = datetime.datetime.now().strftime('%Y%m%d_%H%M%S')
    nome_arquivo = os.path.basename(path_arquivo).replace('.xlsx', '')
    caminho_backup = os.path.join(pasta_backup, f"{nome_arquivo}_backup_{timestamp}.xlsx")
    shutil.copy(path_arquivo, caminho_backup)

    #Verifica quantos backups existem remove os mais antigos mantendo apenas os mais recentes
    backups_existentes = sorted([f for f in os.listdir(pasta_backup) if f.startswith(nome_arquivo)], reverse=True)

    if len(backups_existentes) > max_backups:
        for backup_antigo in backups_existentes[max_backups:]:
            os.remove(os.path.join(pasta_backup, backup_antigo))



def atualizar_abas_centros(path_arquivo):
    df_geral = carregar_planilha(path_arquivo)
    centros_custo = df_geral['CENTRO'].unique()

    for centro in centros_custo:
        df_filtrado = filtrar_por_centro(df_geral, centro)
        atualizar_ou_criar_aba(path_arquivo, df_filtrado, centro)
def validar_campos(novo_dado):
    erros = []


    if not novo_dado['DATA']:
        erros.append("Data não informada.")

    if novo_dado['VALOR'] is None or novo_dado['VALOR'] <= 0:
        erros.append("Valor deve ser maior que zero.")
    
    if novo_dado['VALOR'] == '':
        erros.append("Valor não informado.")
    
    if not novo_dado['FORNECEDOR']:
        erros.append("Fornecedor não informado.")

    if not novo_dado['CENTRO']:
        erros.append("Centro de Custo não informado.")
    
    return erros

def exibir_erro(erros):
    mensagem = "\n".join(erros)
    messagebox.showerror("Erro no preenchimento", f"Corrija os seguintes erros:\n\n{mensagem}")

def converter_valor(valor_texto):
    """Converte o texto do valor em float, aceitando formatos diferentes."""
    try:
        # Remover espaços em branco e converter para float
        valor_numero = float(valor_texto)
        return valor_numero
    except ValueError:
        return None  # Retorna None se a conversão falhar

def adicionar_dados():
    #Capturar os dados do formulário e convertê-los para caixa alta, exceto o valor
    try:
        valor_texto = entry_valor.get().strip()
        if valor_texto:
            valor = converter_valor(valor_texto)
            if valor is None:
                raise ValueError("O campo Valor deve conter um número válido.")
            
            novo_dado = {
        'DATA': entry_data.get(),
        'VALOR': valor, 
        'FORNECEDOR': combobox_fornecedor.get().upper(),
        'DESCRIÇÃO': combobox_descricao.get().upper(),
        'CENTRO': combobox_centro.get().upper(),
        'OBSERVAÇÃO': entry_observacao.get().upper(),
        'DADOS': entry_dados.get().upper()
        }
            
        erros = validar_campos(novo_dado) #valida os dados antes de continuar
        if erros:
            exibir_erro(erros)
            return #finaliza a execucao em caso de erros
        
        sucesso=adicionar_dados_financeiro_geral(PATH_ARQUIVO, novo_dado)

        if sucesso:
            messagebox.showinfo("Sucesso", "Dados adicionados e abas atualizadas com sucesso!")
            atualizar_opcoes()
            reiniciar_formulario()  #Reinicia os campos de entrada
    
    except ValueError:
        messagebox.showerror("Erro no preenchimento", "O campo Valor deve conter um número válido.")
        
    except Exception as e:
        messagebox.showerror("Erro", f"Ocorreu um erro: {e}")

def atualizar_opcoes():
    global fornecedores, centros 
    fornecedores = carregar_opcoes(PATH_ARQUIVO, 'FORNECEDOR')
    centros = carregar_opcoes(PATH_ARQUIVO, 'CENTRO')

    combobox_fornecedor['values'] = fornecedores
    combobox_centro['values'] = centros

def formatar_data(event):
    conteudo = entry_data.get()
    if len(conteudo) == 2 or len(conteudo) == 5:
        entry_data.insert(tk.END, '/')  #Insere uma '/' após o mês e o dia



def reiniciar_formulario():
    #Limpa todos os campos de entrada após a inserção dos dados
    entry_data.delete(0, tk.END)
    entry_valor.delete(0, tk.END)
    combobox_fornecedor.delete(0, tk.END)
    combobox_descricao.delete(0, tk.END)
    combobox_centro.delete(0, tk.END)
    entry_observacao.delete(0, tk.END)
    entry_dados.delete(0, tk.END)

#Filtra as opções de fornecedor com base no que foi digitado (carrega os dados cadastrados (fornecedores ou centros))
def carregar_opcoes(path_arquivo, coluna):
    df= pd.read_excel(path_arquivo, sheet_name='Financeiro Geral')
    return df[coluna].dropna().unique().tolist() #remove valores NaN e retorna uma lista única

#Função de auto-preenchimento para combobox
def combobox_autocomplete(event, combobox, opcoes):
    #Pega o conteúdo do campo de entrada
    texto = combobox.get()
    #Filtra as opções com base no conteúdo do campo de entrada
    opcoes_filtradas = [opcao for opcao in opcoes if opcao.lower().startswith(texto.lower())]
    #Preenche as opções filtradas na combobox
    combobox['values'] = opcoes_filtradas
    if opcoes_filtradas:
        combobox.event_generate('<Down>')#abre a lista suspensa

#carrega as options de fornecedores e centros
fornecedores = carregar_opcoes(PATH_ARQUIVO, 'FORNECEDOR')
centros = carregar_opcoes(PATH_ARQUIVO, 'CENTRO')

#Função para melhorar a aparência da interface
def configurar_interface():
    root.configure(bg='#f2f2f2')
    for widget in root.winfo_children():
        widget.grid_configure(padx=10, pady=5)

def finalizar_aplicacao():
    resposta = messagebox.askyesno("Finalizar", "Você tem certeza que deseja finalizar a aplicação?")
    if resposta:
        root.destroy()#Fecha a janela principal
        sys.exit()

def centralizar_janela():
    largura_janela = 350  
    altura_janela = 300   

    # Obtendo a largura e altura da tela
    largura_tela = root.winfo_screenwidth()
    altura_tela = root.winfo_screenheight()

    # Calculando a posição x e y para centralizar a janela
    pos_x = (largura_tela // 2) - (largura_janela // 2)
    pos_y = (altura_tela // 2) - (altura_janela // 2)

    # Definindo a geometria da janela
    root.geometry(f"{largura_janela}x{altura_janela}+{pos_x}+{pos_y}")
    
#Criação da interface gráfica
root = tk.Tk()
root.title("Adicionar Dados Financeiros")

#Centraliza a janela
centralizar_janela()

#Configurar cores e espaçamento
root.configure(bg='#f2f2f2')

#Fontes e estilos
font_label = ("Arial", 10, "bold")
font_entry = ("Arial", 10)


#Elementos da interface gráfica
tk.Label(root, text="Data (MM/DD/AAAA):", font=font_label, bg='#f2f2f2').grid(row=0, column=0, sticky="e")
entry_data = DateEntry(root, font=font_entry, date_pattern="mm/dd/yyyy")
entry_data.grid(row=0, column=1)

tk.Label(root, text="Valor (R$):", font=font_label, bg='#f2f2f2').grid(row=1, column=0, sticky="e")
entry_valor = tk.Entry(root, font=font_entry)
entry_valor.grid(row=1, column=1)

tk.Label(root, text="Fornecedor:", font=font_label, bg='#f2f2f2').grid(row=2, column=0, sticky="e")
combobox_fornecedor = ttk.Combobox(root, font=font_entry)
combobox_fornecedor.grid(row=2, column=1)
combobox_fornecedor['values'] = fornecedores
combobox_fornecedor.bind("<KeyRelease>", lambda event: combobox_autocomplete(event, combobox_fornecedor, fornecedores))

descricoes = ['Entrada', 'Saída']

tk.Label(root, text="Descrição", font=font_label, bg='#f2f2f2').grid(row=3, column=0, sticky="e")
combobox_descricao = ttk.Combobox(root, font=font_entry)
combobox_descricao.grid(row=3, column=1)
combobox_descricao['values'] = descricoes #Adiciona as opções

tk.Label(root, text="Centro de Custo:", font=font_label, bg='#f2f2f2').grid(row=4, column=0, sticky="e")
combobox_centro = ttk.Combobox(root, font=font_label)
combobox_centro.grid(row=4, column=1)
combobox_centro['values'] = centros
combobox_centro.bind("<KeyRelease>", lambda event: combobox_autocomplete(event, combobox_centro, centros))

tk.Label(root, text="Observação:", font=font_label, bg='#f2f2f2').grid(row=5, column=0, sticky="e")
entry_observacao = tk.Entry(root, font=font_entry)
entry_observacao.grid(row=5, column=1)

tk.Label(root, text="Dados:", font=font_label, bg='#f2f2f2').grid(row=6, column=0, sticky="e")
entry_dados = tk.Entry(root, font=font_entry)
entry_dados.grid(row=6, column=1)

#Botão de Adicionar Dados
btn_adicionar = tk.Button(root, text="Adicionar Dados", font=("Arial", 12), bg='#4CAF50', fg='white', command=adicionar_dados)
btn_adicionar.grid(row=7, column=0, pady=10, sticky="ew")  # Ajuste o sticky para 'ew' para expandir

#Botão de Finalizar
btn_finalizar = tk.Button(root, text="Finalizar", font=("Arial", 12), bg='#f44336', fg='white', command=finalizar_aplicacao)
btn_finalizar.grid(row=7, column=1, pady=10, padx=(10, 0))  # Adicione um espaço com padx

#Aplicar configurações de aparência
configurar_interface()

root.mainloop()